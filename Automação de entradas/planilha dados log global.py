import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import numbers

def ler_e_analisar_log(caminho_arquivo_log):
    with open(caminho_arquivo_log, 'r') as f:
        log_data = f.read()
    
    resultados_re = re.compile(r'Ultimos 3 resultados: (\w+), (\w+), (\w+)')
    porcentagens_25_re = re.compile(r'Ultimas 25 porcentagens: ([\d.]+), ([\d.]+), ([\d.]+)')
    porcentagens_50_re = re.compile(r'Ultimas 50 porcentagens: ([\d.]+), ([\d.]+), ([\d.]+)')
    porcentagens_100_re = re.compile(r'Ultimas 100 porcentagens: ([\d.]+), ([\d.]+), ([\d.]+)')
    porcentagens_500_re = re.compile(r'Ultimas 500 porcentagens: ([\d.]+), ([\d.]+), ([\d.]+)')
    
    resultados = resultados_re.findall(log_data)
    porcentagens_25 = porcentagens_25_re.findall(log_data)
    porcentagens_50 = porcentagens_50_re.findall(log_data)
    porcentagens_100 = porcentagens_100_re.findall(log_data)
    porcentagens_500 = porcentagens_500_re.findall(log_data)
    
    return resultados, porcentagens_25, porcentagens_50, porcentagens_100, porcentagens_500

def comparar_percentuais(atual, oposta):
    if atual > oposta:
        return '>'
    elif atual < oposta:
        return '<'
    else:
        return '='

def analisar_padroes(resultados, porcentagens_25, porcentagens_50, porcentagens_100, porcentagens_500):
    padroes_analise = defaultdict(lambda: {'acertos': 0, 'acertos_branco': 0, 'erros': 0, 'jogadas': 0, 'erros_consecutivos': 0, 'max_erros_consecutivos': 0})

    for i in range(len(resultados) - 2):
        cor_atual = resultados[i][0]
        if resultados[i][0] == resultados[i][1] == resultados[i][2]:
            if cor_atual == "black":
                cor_oposta = "red"
            elif cor_atual == "red":
                cor_oposta = "black"
            else:
                continue

            try:
                porcentagem_atual_25 = float(porcentagens_25[i][2] if cor_atual == "red" else porcentagens_25[i][1])
                porcentagem_oposta_25 = float(porcentagens_25[i][1] if cor_atual == "red" else porcentagens_25[i][2])
                porcentagem_atual_50 = float(porcentagens_50[i][2] if cor_atual == "red" else porcentagens_50[i][1])
                porcentagem_oposta_50 = float(porcentagens_50[i][1] if cor_atual == "red" else porcentagens_50[i][2])
                porcentagem_atual_100 = float(porcentagens_100[i][2] if cor_atual == "red" else porcentagens_100[i][1])
                porcentagem_oposta_100 = float(porcentagens_100[i][1] if cor_atual == "red" else porcentagens_100[i][2])
                porcentagem_atual_500 = float(porcentagens_500[i][2] if cor_atual == "red" else porcentagens_500[i][1])
                porcentagem_oposta_500 = float(porcentagens_500[i][1] if cor_atual == "red" else porcentagens_500[i][2])
            except IndexError:
                continue

            comparacao_25 = comparar_percentuais(porcentagem_atual_25, porcentagem_oposta_25)
            comparacao_50 = comparar_percentuais(porcentagem_atual_50, porcentagem_oposta_50)
            comparacao_100 = comparar_percentuais(porcentagem_atual_100, porcentagem_oposta_100)
            comparacao_500 = comparar_percentuais(porcentagem_atual_500, porcentagem_oposta_500)

            chave_padrao = (porcentagem_atual_25, comparacao_25, comparacao_50, comparacao_100, comparacao_500)
            acerto = resultados[i + 1][0] != cor_atual

            padroes_analise[chave_padrao]['jogadas'] += 1

            if acerto:
                if padroes_analise[chave_padrao]['erros_consecutivos'] > padroes_analise[chave_padrao]['max_erros_consecutivos']:
                    padroes_analise[chave_padrao]['max_erros_consecutivos'] = padroes_analise[chave_padrao]['erros_consecutivos']
                padroes_analise[chave_padrao]['erros_consecutivos'] = 0
                if cor_atual == "white":
                    padroes_analise[chave_padrao]['acertos_branco'] += 1
                else:
                    padroes_analise[chave_padrao]['acertos'] += 1
            else:
                padroes_analise[chave_padrao]['erros'] += 1
                padroes_analise[chave_padrao]['erros_consecutivos'] += 1

                if padroes_analise[chave_padrao]['erros_consecutivos'] > padroes_analise[chave_padrao]['max_erros_consecutivos']:
                    padroes_analise[chave_padrao]['max_erros_consecutivos'] = padroes_analise[chave_padrao]['erros_consecutivos']

        elif resultados[i][0] == "white" and resultados[i][1] == "black" and resultados[i][2] == "black":
            try:
                porcentagem_atual_25 = float(porcentagens_25[i][2])
                porcentagem_oposta_25 = float(porcentagens_25[i][1])
                porcentagem_atual_50 = float(porcentagens_50[i][2])
                porcentagem_oposta_50 = float(porcentagens_50[i][1])
                porcentagem_atual_100 = float(porcentagens_100[i][2])
                porcentagem_oposta_100 = float(porcentagens_100[i][1])
                porcentagem_atual_500 = float(porcentagens_500[i][2])
                porcentagem_oposta_500 = float(porcentagens_500[i][1])
            except IndexError:
                continue

            comparacao_25 = comparar_percentuais(porcentagem_atual_25, porcentagem_oposta_25)
            comparacao_50 = comparar_percentuais(porcentagem_atual_50, porcentagem_oposta_50)
            comparacao_100 = comparar_percentuais(porcentagem_atual_100, porcentagem_oposta_100)
            comparacao_500 = comparar_percentuais(porcentagem_atual_500, porcentagem_oposta_500)

            chave_padrao = (porcentagem_atual_25, comparacao_25, comparacao_50, comparacao_100, comparacao_500)
            acerto = resultados[i + 1][0] != "white"

            padroes_analise[chave_padrao]['jogadas'] += 1

            if acerto:
                if padroes_analise[chave_padrao]['erros_consecutivos'] > padroes_analise[chave_padrao]['max_erros_consecutivos']:
                    padroes_analise[chave_padrao]['max_erros_consecutivos'] = padroes_analise[chave_padrao]['erros_consecutivos']
                padroes_analise[chave_padrao]['erros_consecutivos'] = 0
                padroes_analise[chave_padrao]['acertos_branco'] += 1
            else:
                padroes_analise[chave_padrao]['erros'] += 1
                padroes_analise[chave_padrao]['erros_consecutivos'] += 1

                if padroes_analise[chave_padrao]['erros_consecutivos'] > padroes_analise[chave_padrao]['max_erros_consecutivos']:
                    padroes_analise[chave_padrao]['max_erros_consecutivos'] = padroes_analise[chave_padrao]['erros_consecutivos']

        elif resultados[i][0] == "white" and resultados[i][1] == "red" and resultados[i][2] == "red":
            try:
                porcentagem_atual_25 = float(porcentagens_25[i][2])
                porcentagem_oposta_25 = float(porcentagens_25[i][1])
                porcentagem_atual_50 = float(porcentagens_50[i][2])
                porcentagem_oposta_50 = float(porcentagens_50[i][1])
                porcentagem_atual_100 = float(porcentagens_100[i][2])
                porcentagem_oposta_100 = float(porcentagens_100[i][1])
                porcentagem_atual_500 = float(porcentagens_500[i][2])
                porcentagem_oposta_500 = float(porcentagens_500[i][1])
            except IndexError:
                continue

            comparacao_25 = comparar_percentuais(porcentagem_atual_25, porcentagem_oposta_25)
            comparacao_50 = comparar_percentuais(porcentagem_atual_50, porcentagem_oposta_50)
            comparacao_100 = comparar_percentuais(porcentagem_atual_100, porcentagem_oposta_100)
            comparacao_500 = comparar_percentuais(porcentagem_atual_500, porcentagem_oposta_500)

            chave_padrao = (porcentagem_atual_25, comparacao_25, comparacao_50, comparacao_100, comparacao_500)
            acerto = resultados[i + 1][0] != "white"

            padroes_analise[chave_padrao]['jogadas'] += 1

            if acerto:
                if padroes_analise[chave_padrao]['erros_consecutivos'] > padroes_analise[chave_padrao]['max_erros_consecutivos']:
                    padroes_analise[chave_padrao]['max_erros_consecutivos'] = padroes_analise[chave_padrao]['erros_consecutivos']
                padroes_analise[chave_padrao]['erros_consecutivos'] = 0
                padroes_analise[chave_padrao]['acertos_branco'] += 1
            else:
                padroes_analise[chave_padrao]['erros'] += 1
                padroes_analise[chave_padrao]['erros_consecutivos'] += 1

                if padroes_analise[chave_padrao]['erros_consecutivos'] > padroes_analise[chave_padrao]['max_erros_consecutivos']:
                    padroes_analise[chave_padrao]['max_erros_consecutivos'] = padroes_analise[chave_padrao]['erros_consecutivos']

    return padroes_analise

def calcular_assertividade_acertos(padroes_analise):
    for chave, dados in padroes_analise.items():
        acertos = dados['acertos']
        acertos_branco = dados['acertos_branco']
        jogadas = dados['jogadas']

        assertividade_acertos = (acertos / jogadas) * 100 if jogadas > 0 else 0
        assertividade_acertos_branco = (acertos_branco / jogadas) * 100 if jogadas > 0 else 0
        assertividade_acertos_total = ((acertos + acertos_branco) / jogadas) * 100 if jogadas > 0 else 0

        dados['assertividade_acertos'] = assertividade_acertos
        dados['assertividade_acertos_branco'] = assertividade_acertos_branco
        dados['assertividade_acertos_total'] = assertividade_acertos_total


def gerar_planilha_excel(padroes_analise, caminho_arquivo_excel):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Análise de Comparações'

    header = ["Percentual", "Comp 25", "Comp 50", "Comp 100", "Comp 500", "Acertos", "Brancos", "Erros", "Jogadas", "Últimos Erros Consecutivos", "Máx. Erros Consecutivos", "Assertividade Acertos (%)", "Assertividade Branco (%)", "Assertividade Total (%)"]
    sheet.append(header)

    for chave_padrao, dados in padroes_analise.items():
        percentual_atual, comp_25, comp_50, comp_100, comp_500 = chave_padrao
        acertos = dados['acertos']
        acertos_branco = dados['acertos_branco']
        erros = dados['erros']
        jogadas = dados['jogadas']
        erros_consecutivos = dados['erros_consecutivos']
        max_erros_consecutivos = dados['max_erros_consecutivos']
        assertividade_acertos = dados['assertividade_acertos']
        assertividade_acertos_branco = dados['assertividade_acertos_branco']
        assertividade_acertos_total = dados['assertividade_acertos_total']

        row = [percentual_atual, comp_25, comp_50, comp_100, comp_500, acertos, acertos_branco, erros, jogadas, erros_consecutivos, max_erros_consecutivos, assertividade_acertos, assertividade_acertos_branco, assertividade_acertos_total]
        sheet.append(row)

    # Ajuste para formatar as colunas de porcentagem como números
    for col in range(2, 6):
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

    workbook.save(caminho_arquivo_excel)
    print(f"Planilha Excel gerada com sucesso em: {caminho_arquivo_excel}")

def main():
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    arquivo_log = os.path.join(desktop_path, 'LOGS', 'log global.txt')
    caminho_arquivo_excel = os.path.join(desktop_path, 'LOGS', 'Dados log global.xlsx')
    
    resultados, porcentagens_25, porcentagens_50, porcentagens_100, porcentagens_500 = ler_e_analisar_log(arquivo_log)
    padroes_analise = analisar_padroes(resultados, porcentagens_25, porcentagens_50, porcentagens_100, porcentagens_500)
    calcular_assertividade_acertos(padroes_analise)
    gerar_planilha_excel(padroes_analise, caminho_arquivo_excel)

if __name__ == "__main__":
    main()
